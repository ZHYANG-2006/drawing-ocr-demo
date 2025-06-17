# uploader/extract_utils.py
import fitz  # PyMuPDF
import io

import openpyxl
from openpyxl.styles import Font
from datetime import datetime
from openpyxl.workbook import Workbook
from django.http import JsonResponse

def extract_text_from_pdf(pdf_file_bytes):
    # 确保传递给 fitz.open() 的是字节流
    doc = fitz.open("pdf", pdf_file_bytes)  # 使用 "pdf" 表示它是一个 PDF 字节流

    all_text = ""
    for page in doc:
        all_text += page.get_text()  # 提取每一页的文本
    return all_text


import re


def process_extracted_data(extracted_text):
    # 只提取 "检索结果为:" 和 "日期" 到 "比较" 的部分
    result = []

    # 提取检索结果（例如 "D14174-TP1A-R 20240904-0465 3M93020LE K00004197 DR-3"）
    result_match = re.search(r"检索结果为:(.*?)日期", extracted_text, re.DOTALL)
    if result_match:
        result.append(result_match.group(1).strip())
        # 提取日期
        date_match = re.search(r"日期:(.*?)检索算法", extracted_text, re.DOTALL)
        if date_match:
            result.append(date_match.group(1).strip())
        cmp_match = re.search(r"比较:(.*?)\n", extracted_text, re.DOTALL)
        if cmp_match:
            result.append(cmp_match.group(1).strip())
    else:
        # 提取数值数据（例如波数和吸光度）
        # data_match = re.findall(r"\d+\.\d+", extracted_text)
        # result.extend(data_match)  # 将所有数字数据加入结果
        lines = extracted_text.splitlines()

        pattern = r"[A-Za-z0-9\s-]*[A-Za-z][0-9]+[A-Za-z0-9\s-]*"
        # 查找包含字母和数字的字符串，并提取它之前的数据
        for i, line in enumerate(lines):
            # 如果找到了包含字母和数字的目标字符串
            if re.search(pattern, line):
                # 提取目标字符串及它前面的每一行数据
                result = lines[:i + 1]  # 取到当前行
                break
    return result


import openpyxl

month_map = {
    '10月': 'Oct', '11月': 'Nov', '12月': 'Dec', '03月': 'Mar', '04月': 'Apr', '05月': 'May', '06月': 'Jun',
    '07月': 'Jul', '08月': 'Aug', '09月': 'Sep', '01月': 'Jan', '02月': 'Feb', '3月': 'Mar', '4月': 'Apr', '5月': 'May', '6月': 'Jun',
    '7月': 'Jul', '8月': 'Aug', '9月': 'Sep', '1月': 'Jan', '2月': 'Feb'
}

def save_data_to_excel(filtered_data, extracted_text):
    # 创建一个新的 Excel 文件对象
    wb = Workbook()
    ws = wb.active
    if '被检索的区域' in extracted_text:
        params = filtered_data[0].split()
        params_date = filtered_data[1].split('\n')
        clean_date_str = params_date[0].split(' ', 1)[1].split(' (GMT')[0]

        for cn_month, en_month in month_map.items():
            clean_date_str = clean_date_str.replace(cn_month, en_month)
        # 解析日期并格式化为目标格式
        date_obj = datetime.strptime(clean_date_str, "%b %d %H:%M:%S %Y")

        # 格式化为目标格式：2024/9/5 02:44:53
        formatted_date = date_obj.strftime("%Y/%m/%d %H:%M:%S")
        # return render(request, 'upload_pdf.html',
        #               {'form': PDFUploadForm(), 'ID': 1, 'Incoming Date': params_date[0], '流水号': params[1],
        #                'P/N': params[0], '原胶型号': params[2], '供应商': params[4], '原胶D/C': params[3],
        #                '匹配率': extracted_data[2]})
        ws.append(['ID', 'Incoming Date', '流水号', 'P/N', '原胶型号', '供应商', '原胶D/C', '匹配率'])
        if len(params) >= 5:
            ws.append([1, formatted_date, params[1], params[0], params[2], params[4], params[3], filtered_data[2]])
        elif len(params) == 4:
            ws.append([1, formatted_date, params[1], params[0], params[2], params[3], '', filtered_data[2]])
        else:
            raise ValueError(f"Unexpected params length: {len(params)} - {params}")
    else:
        # 将数据分为头部和详情
        if len(filtered_data) > 0:
            extract_head = filtered_data[-1]  # 最后一个元素作为头部
            detail_data = filtered_data[:-1]  # 其余元素作为详情
            float_array = [float(item) for item in detail_data]
            sorted_data = sorted(float_array)
            ws.append([extract_head])
            for row in sorted_data:
                ws.append([row])

    # 创建一个 BytesIO 流对象，用于将 Excel 文件保存在内存中
    excel_file_stream = io.BytesIO()

    # 将 Excel 文件保存到 BytesIO 流中
    wb.save(excel_file_stream)

    # 重置流的位置到开头，准备读取
    excel_file_stream.seek(0)
    return excel_file_stream

# 计数字符串中的数字数量
def count_digits(s):
    return len(re.findall(r'\d', s))

def is_valid_pn_format(val: str) -> bool:
    """
    判断字符串是否符合 pn 的要求：
    1. 包含正好两个 '-'。
    2. 拆分后有 3 部分 (parts)。
    3. 第一部分以字母开头，后面只能是字母或数字。
    """
    if val.count('-') >= 2:
        parts = val.split('-')
        if len(parts) >= 3:
            # 使用正则来判断第一部分是否符合以字母开头，后面只能是字母或数字
            return re.match(r'^[A-Za-z][A-Za-z0-9]*$', parts[0]) is not None
    return False

def process_file(file):
    file.seek(0)
    # 读取上传的 PDF 文件到内存
    pdf_file_bytes = file.read()

    # 使用内存中的文件进行提取
    extracted_text = extract_text_from_pdf(pdf_file_bytes)
    extracted_data = process_extracted_data(extracted_text)  # 处理提取的数据
    if '被检索的区域' in extracted_text:
        params = extracted_data[0].split()
        if len(params) > 5:
            params[4] = params[4] + " " + params[5]
        params_date = extracted_data[1].split('\n')
        clean_date_str = params_date[0].split(' ', 1)[1].split(' (GMT')[0]

        for cn_month, en_month in month_map.items():
            clean_date_str = clean_date_str.replace(cn_month, en_month)
        # 解析日期并格式化为目标格式
        date_obj = datetime.strptime(clean_date_str, "%b %d %H:%M:%S %Y")
        supplier = ''
        original_rubber_dc = ''
        if len(params) >= 5:
            supplier = params[4]
            original_rubber_dc = params[3]
        elif len(params) == 4:
            supplier = params[3]
            original_rubber_dc = ''
        elif len(params) == 3:
            supplier = ''
            original_rubber_dc = ''
        else:
            raise ValueError(f"Unexpected params length: {len(params)} - {params}")
        context = {
            'ID': 1,
            'incoming_date': date_obj,
            'supplier': supplier,
            'original_rubber_dc': original_rubber_dc,
            'match_rate': extracted_data[2]
            # ...
        }
        # 处理params中的值
        params_0_digits = count_digits(params[0])
        params_1_digits = count_digits(params[1])
        params_2_digits = count_digits(params[2])

        # 逻辑处理：如果包含两个"-"的值赋给pn
        if is_valid_pn_format(params[0]):
            context['pn'] = params[0]
            params[0] = None
        elif is_valid_pn_format(params[1]):
            context['pn'] = params[1]
            params[1] = None
        elif is_valid_pn_format(params[2]):
            context['pn'] = params[2]
            params[2] = None

        # 处理数字数量大于等于10的值赋给verification_serial_number
        if params_1_digits >= 10:
            context['verification_serial_number'] = params[1]
            params[1] = None
        elif params_0_digits >= 10:
            context['verification_serial_number'] = params[0]
            params[0] = None
        elif params_2_digits >= 10:
            context['verification_serial_number'] = params[2]
            params[2] = None

        # 剩下的值赋给original_rubber_model
        if params[0]:
            context['original_rubber_model'] = params[0]
        elif params[1]:
            context['original_rubber_model'] = params[1]
        elif params[2]:
            context['original_rubber_model'] = params[2]
        # context = {
        #     'ID': 1,
        #     'incoming_date': date_obj,
        #     'verification_serial_number': params[1],
        #     'pn': params[0],
        #     'original_rubber_model': params[2],
        #     'supplier': params[4],
        #     'original_rubber_dc': params[3],
        #     'match_rate': extracted_data[2]
        #     # ...
        # }
        # context = {
        #     'ID': 1,
        #     'Incoming Date': params_date[0],
        #     '流水号': params[1],
        #     'P/N': params[2],
        #     '原胶型号': params[2],
        #     '供应商': params[4],
        #     '原胶D/C': params[3],
        #     '匹配率': extracted_data[2]
        #     # ...
        # }
        return context
        # 格式化为目标格式：2024/9/5 02:44:53
        formatted_date = date_obj.strftime("%Y/%m/%d %H:%M:%S")
        # return render(request, 'upload_pdf.html',
        #               {'form': PDFUploadForm(), 'ID': 1, 'Incoming Date': params_date[0], '流水号': params[1],
        #                'P/N': params[0], '原胶型号': params[2], '供应商': params[4], '原胶D/C': params[3],
        #                '匹配率': extracted_data[2]})
        # return render(request, 'upload_pdf.html',
        #               {'form': PDFUploadForm(), 'ID': 1, 'p1': formatted_date, 'p2': params[1], 'p3': params[0],
        #                'p4': params[2], 'p5': params[4], 'p6': params[3], 'p7': extracted_data[2]})
    else:
        # 将数据分为头部和详情
        if len(extracted_data) > 0:
            extract_head = extracted_data[-1]  # 最后一个元素作为头部
            detail_data = extracted_data[:-1]  # 其余元素作为详情
            float_array = [float(item) for item in detail_data]
            params = extract_head.split()

            # context = {
            #     'verification_serial_number': params[0],
            #     'rubber_model': params[1],
            #     'material_number': params[2],
            #     'original_rubber_lot': params[3],
            #     'supplier': params[4],
            #     'measurement_count': params[5],
            #     'detail_data': sorted(float_array),
            #     # ...
            # }
            context = {
                'original_rubber_lot': params[3],
                'supplier': params[4],
                'measurement_count': params[5],
                'detail_data': sorted(float_array),
                # ...
            }
            params_0_digits = count_digits(params[0])
            params_1_digits = count_digits(params[1])
            params_2_digits = count_digits(params[2])

            if is_valid_pn_format(params[0]):
                context['material_number'] = params[0]
                params[0] = None
            elif is_valid_pn_format(params[1]):
                context['material_number'] = params[1]
                params[1] = None
            elif is_valid_pn_format(params[2]):
                context['material_number'] = params[2]
                params[2] = None
            if params_1_digits >= 10:
                context['verification_serial_number'] = params[1]
                params[1] = None
            elif params_0_digits >= 10:
                context['verification_serial_number'] = params[0]
                params[0] = None
            elif params_2_digits >= 10:
                context['verification_serial_number'] = params[2]
                params[2] = None
            if params[0]:
                context['rubber_model'] = params[0]
            elif params[1]:
                context['rubber_model'] = params[1]
            elif params[2]:
                context['rubber_model'] = params[2]
            return context
        # return render(request, 'upload_pdf.html',
        #               {'form': PDFUploadForm(), 'extract_head': extract_head, 'detail_data': sorted(float_array)})

def export_excel(context, type_pdf):
    # 1. 创建工作簿 & 工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    if type_pdf == 'peak':
        # 2. 找到最大峰值数量
        max_peaks = 0
        for item in context:
            peaks_len = len(item.get("peaks", []))
            if peaks_len > max_peaks:
                max_peaks = peaks_len

        # 3. 构建表头
        # 固定列头
        headers = [
            "检验流水号",
            "胶型号",
            "料号",
            "原胶LOT",
            "供应商",
            "测量次数",
        ]
        # 动态列头，如 ["峰值1", "峰值2", ..., "峰值{max_peaks}"]
        peak_headers = [f"峰值{i+1}" for i in range(max_peaks)]

        # 合并表头
        all_headers = headers + peak_headers

        # 写入表头到第一行
        for col_index, header in enumerate(all_headers, start=1):
            cell = ws.cell(row=1, column=col_index, value=header)
            cell.font = Font(bold=True)  # 加粗

        # 4. 填写数据行
        # 每个 item 对应一行
        current_row = 2
        for item in context:
            # 写入固定列
            ws.cell(row=current_row, column=1, value=item.get("verification_serial_number", ""))
            ws.cell(row=current_row, column=2, value=item.get("rubber_model", ""))
            ws.cell(row=current_row, column=3, value=item.get("material_number", ""))
            ws.cell(row=current_row, column=4, value=item.get("original_rubber_lot", ""))
            ws.cell(row=current_row, column=5, value=item.get("supplier", ""))
            ws.cell(row=current_row, column=6, value=item.get("measurement_count", ""))

            # 写入 峰值
            peaks = item.get("peaks", [])
            for i, peak_dict in enumerate(peaks):
                # i=0 => 峰值1, i=1 => 峰值2
                col_index = 7 + i  # 第7列开始放峰值
                peak_value = peak_dict.get("peak_value", "")
                ws.cell(row=current_row, column=col_index, value=peak_value)

            current_row += 1

        # 5. 返回 Workbook (或直接保存/或转为 HTTP 响应)
        return wb
    else:
        # 3. 构建表头
        # 固定列头
        headers = [
            "Incoming Date",
            "流水号",
            "P/N",
            "原胶型号",
            "供应商",
            "原胶D/C",
            "匹配率",
        ]

        # 写入表头到第一行
        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_index, value=header)
            cell.font = Font(bold=True)  # 加粗

        # 4. 填写数据行
        # 每个 item 对应一行
        current_row = 2
        for item in context:
            # 写入固定列
            ws.cell(row=current_row, column=1, value=item.get("incoming_date", ""))
            ws.cell(row=current_row, column=2, value=item.get("verification_serial_number", ""))
            ws.cell(row=current_row, column=3, value=item.get("pn", ""))
            ws.cell(row=current_row, column=4, value=item.get("original_rubber_model", ""))
            ws.cell(row=current_row, column=5, value=item.get("supplier", ""))
            ws.cell(row=current_row, column=6, value=item.get("original_rubber_dc", ""))
            ws.cell(row=current_row, column=7, value=item.get("match_rate", ""))

            current_row += 1

        # 5. 返回 Workbook (或直接保存/或转为 HTTP 响应)
        return wb